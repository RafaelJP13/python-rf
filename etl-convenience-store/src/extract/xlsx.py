from openpyxl import load_workbook

from src.exceptions import ExtractionError
from src.types import Product

FILE_PATH = "data/raw/products.xlsx"

REQUIRED_COLUMNS = {
    "id",
    "name",
    "category",
    "price",
    "supply",
}

def ExtractXlsx() -> list[Product]:
    try:
        workbook = load_workbook(
            FILE_PATH,
            read_only=True,
            data_only=True,
        )

        worksheet = workbook.active
        
        if worksheet is None:
            raise ExtractionError("The XLSX file does not contain a worksheet!")

        rows = worksheet.iter_rows(values_only=True)

        headers = next(rows, None)

        if headers is None:
            raise ExtractionError("The XLSX file does not contain a header!")

        columns = set(headers)

        missing_columns = REQUIRED_COLUMNS - columns

        if missing_columns:
            raise ExtractionError(
                f"Missing required columns: {missing_columns}"
            )

        column_indexes = {
            column: headers.index(column)
            for column in REQUIRED_COLUMNS
        }

        products: list[Product] = []

        for row_number, row in enumerate(rows, start=2):
            try:
                products.append(
                    {
                        "id": str(row[column_indexes["id"]]),
                        "name": str(row[column_indexes["name"]]),
                        "category": str(row[column_indexes["category"]]),
                        "price": str(row[column_indexes["price"]]),
                        "supply": str(row[column_indexes["supply"]]),
                    }
                )
            except (IndexError, TypeError) as exc:
                raise ExtractionError(
                    f"Invalid XLSX structure at row {row_number}"
                ) from exc

        return products

    except FileNotFoundError as exc:
        raise ExtractionError(
            f"Input file not found: {FILE_PATH}"
        ) from exc